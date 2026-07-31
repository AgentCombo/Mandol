import os
import json
import argparse
import glob
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def parse_args():
    parser = argparse.ArgumentParser(description="生成自适应RAG系统结果的Excel统计报告(严格模式)")
    parser.add_argument(
        "--input-dir", 
        type=str, 
        required=True, 
        help="包含JSON结果文件的输入目录路径"
    )
    parser.add_argument(
        "--output-dir", 
        type=str, 
        default=None, 
        help="Excel文件的输出路径，默认为input-dir"
    )
    return parser.parse_args()

def calculate_total_retrieval_time(result_item):
    """Compute total retrieval time."""
    t_hier = result_item.get("hierarchical_retrieval_time", 0.0) or 0.0
    t_graph = result_item.get("graph_retrieval_time", 0.0) or 0.0
    total_time = t_hier + t_graph

    expansion_results = result_item.get("expansion_retrieval_results")
    if expansion_results and isinstance(expansion_results, list):
        for exp_res in expansion_results:
            
            t_exp = exp_res.get("retrieval_time", 0.0) or 0.0
            total_time += t_exp
            
    return total_time

def process_files(input_dir):
    json_files = glob.glob(os.path.join(input_dir, "*.json"))
    
    all_records = []
    
    print(f"在 {input_dir} 中发现 {len(json_files)} 个JSON文件")

    for file_path in json_files:
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                
            sample_info = data.get("sample_info", {})
            sample_id_global = sample_info.get("sample_id", "unknown")
            results = data.get("results", [])
            
            if not isinstance(results, list):
                continue

            for item in results:
                question = item.get("question", "")
                category = item.get("category", 0)
                
                query_expansion_used = item.get("query_expansion_used", False)
                
                eval_scores = item.get("evaluation_scores", {})
                if eval_scores is None:
                    eval_scores = {}
                
                llm_accuracy = eval_scores.get("llm_accuracy", 0.0)
                
                is_strictly_correct = (abs(llm_accuracy - 1.0) < 1e-6)
                
                total_retrieval_time = calculate_total_retrieval_time(item)
                
                record = {
                    "文件名": os.path.basename(file_path),
                    "样本ID": sample_id_global,
                    "问题内容": question,
                    "问题类别": category,
                    "LLM准确率": llm_accuracy,
                    "是否触发扩展": "是" if query_expansion_used else "否",
                    "是否严格正确(Acc=1.0)": "是" if is_strictly_correct else "否",
                    "总检索耗时(秒)": total_retrieval_time,
                    
                    "_raw_expansion": query_expansion_used,
                    "_raw_strict_success": is_strictly_correct
                }
                all_records.append(record)
                
        except Exception as e:
            print(f"处理文件 {file_path} 时出错: {e}")
            continue
            
    return pd.DataFrame(all_records)

def generate_summary(df):
    if df.empty:
        return pd.DataFrame()

    total_queries = len(df)
    
    # Avoid mutating LogRecord fields before other handlers process the record.
    expansion_triggered_count = df['_raw_expansion'].sum()
    expansion_triggered_rate = expansion_triggered_count / total_queries if total_queries > 0 else 0

    # Avoid mutating LogRecord fields before other handlers process the record.
    no_exp_success_count = len(df[(df['_raw_expansion'] == False) & (df['_raw_strict_success'] == True)])
    no_exp_success_rate = no_exp_success_count / total_queries if total_queries > 0 else 0

    # Avoid mutating LogRecord fields before other handlers process the record.
    exp_success_count = len(df[(df['_raw_expansion'] == True) & (df['_raw_strict_success'] == True)])
    exp_success_rate = exp_success_count / total_queries if total_queries > 0 else 0
    
    # Avoid mutating LogRecord fields before other handlers process the record.
    precision_of_expansion = exp_success_count / expansion_triggered_count if expansion_triggered_count > 0 else 0

    avg_retrieval_time = df['总检索耗时(秒)'].mean()
    total_retrieval_time_sum = df['总检索耗时(秒)'].sum()

    summary_data = {
        "统计指标": [
            "查询总数 (Input)",
            "1. 触发查询扩展数量 (量化器判断不足)",
            "   - 触发比例 (触发数 / 总数)",
            "2. 未扩展但回答正确数量 (漏判但正确, Acc=1.0)",
            "   - 未扩展正确率 (未扩展正确 / 总数)",
            "3. 扩展且回答正确数量 (有效修正, Acc=1.0)",
            "   - 扩展修正率 (扩展正确 / 总数)",
            "   - 扩展准确率 (扩展正确 / 触发数)",
            "4. 平均检索耗时 (秒)",
            "   - 总检索耗时 (秒)"
        ],
        "数值": [
            total_queries,
            expansion_triggered_count,
            f"{expansion_triggered_rate:.2%}",
            no_exp_success_count,
            f"{no_exp_success_rate:.2%}",
            exp_success_count,
            f"{exp_success_rate:.2%}",
            f"{precision_of_expansion:.2%}",
            f"{avg_retrieval_time:.4f}",
            f"{total_retrieval_time_sum:.4f}"
        ]
    }
    
    return pd.DataFrame(summary_data)

def style_excel(file_path):
    """Run style excel."""
    wb = load_workbook(file_path)
    
    header_font = Font(name='微软雅黑', size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    content_font = Font(name='微软雅黑', size=10)
    border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))
    alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            if sheet_name == 'Details' and column_cells[0].value == "问题内容":
                ws.column_dimensions[get_column_letter(column_cells[0].column)].width = 50
            else:
                adjusted_width = (length + 2) * 1.3
                adjusted_width = min(adjusted_width, 60)
                ws.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width

        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.font = content_font
                
                if cell.row == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = alignment
                else:
                    if sheet_name == 'Summary' and cell.column == 1:
                         cell.alignment = left_alignment
                    elif sheet_name == 'Details' and cell.column == 3:
                         cell.alignment = left_alignment
                         cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    else:
                         cell.alignment = alignment

    wb.save(file_path)

def main():
    args = parse_args()
    
    input_dir = args.input_dir
    output_dir = args.output_dir if args.output_dir else input_dir
    
    if not os.path.exists(input_dir):
        print(f"错误: 输入目录 {input_dir} 不存在")
        return
        
    if not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)
        
    df_details = process_files(input_dir)
    
    if df_details.empty:
        print("未提取到任何数据，请检查输入目录下的JSON文件结构。")
        return

    df_summary = generate_summary(df_details)
    
    output_columns = [col for col in df_details.columns if not col.startswith('_')]
    df_details_clean = df_details[output_columns]

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"RAG系统严格统计报告_{timestamp}.xlsx"
    output_path = os.path.join(output_dir, output_filename)
    
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_summary.to_excel(writer, sheet_name='Summary', index=False)
            df_details_clean.to_excel(writer, sheet_name='Details', index=False)
        
        style_excel(output_path)
            
        print(f"\n 统计完成！")
        print(f" Excel报告已保存至: {output_path}")
        print("\n=== 摘要预览 ===")
        print(df_summary.to_string(index=False))
        
    except Exception as e:
        print(f"写入或美化Excel文件失败: {e}")

if __name__ == "__main__":
    main()