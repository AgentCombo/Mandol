#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate Excel Analysis Report for Entity-Relation Benchmark Results

This script reads all qa_*.json files from the individual_reports directory
of entity_relation benchmark, aggregates the results, and creates a comprehensive
Excel report.

Designed specifically for entity_relation benchmark JSON structure:
- retrieval_details.total_entities_retrieved
- retrieval_details.entity_types_found
- total_time field
- question_type: factual, etc.

Format aligned with generate_excel_report_basic_cut.py
"""

import json
import os
import re
import statistics
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from datetime import datetime
import argparse


def load_individual_results(individual_reports_dir: str) -> Tuple[List[Dict], Dict]:
    """
    Load all qa_*.json files from the individual_reports directory.
    
    Returns:
        Tuple of (detailed_results list, computed summary dict)
    """
    reports_path = Path(individual_reports_dir)
    
    if not reports_path.exists():
        raise FileNotFoundError(f"目录不存在: {individual_reports_dir}")
    
    # Find all qa_*.json files
    json_files = list(reports_path.glob('qa_*.json'))
    
    if not json_files:
        raise FileNotFoundError(f"未找到qa_*.json文件: {individual_reports_dir}")
    
    # Sort by qa number
    def extract_qa_number(filepath):
        match = re.search(r'qa_(\d+)\.json', filepath.name)
        return int(match.group(1)) if match else 0
    
    json_files = sorted(json_files, key=extract_qa_number)
    
    detailed_results = []
    failed_count = 0
    
    print(f"正在加载 {len(json_files)} 个QA结果文件...")
    
    for json_file in json_files:
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                result = json.load(f)
                detailed_results.append(result)
        except Exception as e:
            print(f" 加载失败: {json_file.name} - {e}")
            failed_count += 1
    
    print(f"成功加载: {len(detailed_results)} 个, 失败: {failed_count} 个")
    
    # Compute summary from detailed results
    summary = compute_summary(detailed_results)
    
    return detailed_results, summary


def get_question_type(result: Dict) -> str:
    """Helper to extract question type from the JSON structure."""
    qtype = result.get('evaluation', {}).get('input_info', {}).get('question_type')
    if not qtype:
        qtype = result.get('question_type', 'unknown')
    return qtype


def compute_summary(detailed_results: List[Dict]) -> Dict:
    """
    Compute summary statistics from detailed results.
    """
    if not detailed_results:
        return {}
    
    total_tests = len(detailed_results)
    
    # Count successful tests
    successful_tests = sum(1 for r in detailed_results if r.get('success', False))
    failed_tests = total_tests - successful_tests
    
    # Collect metrics
    llm_accuracies = []
    f1_scores = []
    retrieval_times = []
    test_times = []
    
    # By question type
    by_question_type = defaultdict(lambda: {
        'count': 0,
        'llm_accuracies': [],
        'f1_scores': [],
        'retrieval_times': []
    })
    
    # Get retrieval config from first result
    retrieval_config = {}
    if detailed_results:
        first_result = detailed_results[0]
        retrieval_details = first_result.get('retrieval_details', {})
        retrieval_config = {
            'retrieval_methods': [retrieval_details.get('method', 'hybrid_rrf_no_expand')],
            'fusion_method': 'RRF',
            'rerank_method': retrieval_details.get('method', 'N/A'),
            'top_k': retrieval_details.get('total_entities_retrieved', 10)
        }
    
    for result in detailed_results:
        # Get evaluation data
        evaluation = result.get('evaluation', {})
        scores = evaluation.get('scores', {})
        
        # Extract question type
        question_type = get_question_type(result)
        
        # LLM accuracy
        llm_acc = scores.get('llm_accuracy')
        if llm_acc is not None:
            llm_accuracies.append(llm_acc)
            by_question_type[question_type]['llm_accuracies'].append(llm_acc)
        
        # F1 score
        f1 = scores.get('token_f1')
        if f1 is not None:
            f1_scores.append(f1)
            by_question_type[question_type]['f1_scores'].append(f1)
        
        # Retrieval time
        ret_details = result.get('retrieval_details', {})
        ret_time = ret_details.get('retrieval_time')
        if ret_time is not None:
            retrieval_times.append(ret_time)
            by_question_type[question_type]['retrieval_times'].append(ret_time)
        
        # Total time (entity_relation specific)
        total_time = result.get('total_time', 0)
        test_times.append(total_time)
        
        # Count
        by_question_type[question_type]['count'] += 1
    
    # Compute averages for each question type
    by_type_summary = {}
    for qtype, data in by_question_type.items():
        by_type_summary[qtype] = {
            'count': data['count'],
            'avg_llm_accuracy': statistics.mean(data['llm_accuracies']) if data['llm_accuracies'] else 0,
            'avg_f1': statistics.mean(data['f1_scores']) if data['f1_scores'] else 0,
            'avg_retrieval_time': statistics.mean(data['retrieval_times']) if data['retrieval_times'] else 0
        }
    
    # Build summary
    summary = {
        'baseline_version': 'entity_relation_no_expand',
        'dataset_size': f'{total_tests} QAs',
        'total_tests': total_tests,
        'successful_tests': successful_tests,
        'failed_tests': failed_tests,
        'overall_llm_accuracy': statistics.mean(llm_accuracies) if llm_accuracies else 0,
        'overall_f1': statistics.mean(f1_scores) if f1_scores else 0,
        'avg_retrieval_time': statistics.mean(retrieval_times) if retrieval_times else 0,
        'total_benchmark_time': sum(test_times),
        'retrieval_config': retrieval_config,
        'by_question_type': by_type_summary
    }
    
    return summary


def safe_str_truncate(value, max_len: int = 50) -> str:
    """Safely convert value to string and truncate if needed."""
    if value is None:
        return 'N/A'
    str_value = str(value)
    if len(str_value) > max_len:
        return str_value[:max_len] + '...'
    return str_value


def get_zep_accuracy(question_type: str, model: str = 'gpt-4o-mini') -> str:
    """
    Get ZEP accuracy for specific question type and model.
    
    Based on:
    - Table 2: LongMemEval_s overall scores
    - Table 3: LongMemEval_s Question Type Breakdown
    """
    zep_data = {
        'gpt-4o-mini': {
            # Overall score from Table 2
            'overall': 0.638,
            # Question type breakdown from Table 3
            'single-session-preference': 0.533,
            'single-session-assistant': 0.750,
            'temporal-reasoning': 0.541,
            'multi-session': 0.474,
            'knowledge-update': 0.744,
            'single-session-user': 0.929
        },
        'gpt-4o': {
            # Overall score from Table 2
            'overall': 0.712,
            # Question type breakdown from Table 3
            'single-session-preference': 0.567,
            'single-session-assistant': 0.804,
            'temporal-reasoning': 0.624,
            'multi-session': 0.579,
            'knowledge-update': 0.833,
            'single-session-user': 0.929
        }
    }
    
    # Get accuracy for the specific question type or overall
    accuracy = zep_data.get(model, {}).get(question_type, None)
    if accuracy is not None:
        return f"{accuracy:.2%}"
    return 'N/A'


def create_excel_report(detailed_results: List[Dict], summary: Dict, output_path: str, zep_model: str = 'gpt-4o-mini'):
    """Create Excel report with multiple analysis sheets."""
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        print("错误: 需要安装 pandas 和 openpyxl")
        print("请运行: pip install pandas openpyxl")
        return False
    
    # Create Excel writer
    writer = pd.ExcelWriter(output_path, engine='openpyxl')
    
    # Prepare results dict for compatibility
    results = {
        'summary': summary,
        'detailed_results': detailed_results
    }
    
    create_summary_sheet_horizontal(results, writer, zep_model)
    
    create_question_type_comparison_sheet(results, writer, zep_model)
    
    create_question_type_comparison_transposed(results, writer, zep_model)
    
    # Sheets 4+: Individual Question Type Details
    create_individual_question_type_sheets(results, writer)
    
    # Token Usage Analysis
    create_token_analysis_sheet(results, writer)
    
    # Cost Estimation
    create_cost_estimation_sheet_standalone(results, writer)
    
    # Retrieval Performance
    create_retrieval_sheet(results, writer)
    
    # Error Analysis
    create_error_analysis_sheet(results, writer)
    
    # Entity Statistics (entity_relation specific)
    create_entity_statistics_sheet(results, writer)
    
    # Save and close
    writer.close()
    
    # Apply formatting
    apply_excel_formatting(output_path)
    
    return True


def create_summary_sheet_horizontal(results: Dict, writer, zep_model: str = 'gpt-4o-mini'):
    """Create overall summary sheet in horizontal layout with ZEP comparison."""
    import pandas as pd
    
    summary = results.get('summary', {})
    
    zep_overall = get_zep_accuracy('overall', zep_model)
    
    ret_methods = summary.get('retrieval_config', {}).get('retrieval_methods', [])
    if isinstance(ret_methods, list):
        ret_methods_str = ', '.join(str(x) for x in ret_methods)
    else:
        ret_methods_str = str(ret_methods)
    
    data = {
        '基线版本': [summary.get('baseline_version', 'N/A')],
        '数据集': [summary.get('dataset_size', 'N/A')],
        '总测试数': [summary.get('total_tests', 0)],
        '成功数': [summary.get('successful_tests', 0)],
        '失败数': [summary.get('failed_tests', 0)],
        'LLM准确率': [f"{summary.get('overall_llm_accuracy', 0):.2%}"],
        f'ZEP准确率({zep_model})': [zep_overall],
        '平均F1': [f"{summary.get('overall_f1', 0):.4f}"],
        '平均检索时间(秒)': [f"{summary.get('avg_retrieval_time', 0):.2f}"],
        '总耗时(秒)': [f"{summary.get('total_benchmark_time', 0):.2f}"],
        '检索方法': [ret_methods_str],
        '融合方法': [summary.get('retrieval_config', {}).get('fusion_method', 'N/A')],
        '重排方法': [summary.get('retrieval_config', {}).get('rerank_method', 'N/A')],
        'Top-K': [summary.get('retrieval_config', {}).get('top_k', 'N/A')]
    }
    
    df = pd.DataFrame(data)
    df.to_excel(writer, sheet_name='总体概览', index=False)


def create_question_type_comparison_sheet(results: Dict, writer, zep_model: str = 'gpt-4o-mini'):
    """Create question type comparison sheet (ZEP style) with ZEP accuracy."""
    import pandas as pd
    
    by_type = results.get('summary', {}).get('by_question_type', {})
    
    question_types = sorted(by_type.keys())
    
    data = {
        '指标': ['样本数', 'LLM准确率', f'ZEP准确率({zep_model})', 'F1分数', '平均检索时间(秒)']
    }
    
    for qtype in question_types:
        stats = by_type[qtype]
        data[qtype] = [
            stats.get('count', 0),
            f"{stats.get('avg_llm_accuracy', 0):.2%}",
            get_zep_accuracy(qtype, zep_model),
            f"{stats.get('avg_f1', 0):.4f}",
            f"{stats.get('avg_retrieval_time', 0):.2f}"
        ]
    
    df = pd.DataFrame(data)
    df.to_excel(writer, sheet_name='问题类型对比', index=False)


def create_question_type_comparison_transposed(results: Dict, writer, zep_model: str = 'gpt-4o-mini'):
    """Build question type comparison transposed."""
    import pandas as pd
    
    by_type = results.get('summary', {}).get('by_question_type', {})
    
    
    question_types = sorted(by_type.keys())
    
    data = []
    
    for qtype in question_types:
        stats = by_type[qtype]
        
        zep_acc = get_zep_accuracy(qtype, zep_model)
        
        row = {
            '问题类型': qtype,
            '样本数': stats.get('count', 0),
            'LLM准确率': f"{stats.get('avg_llm_accuracy', 0):.2%}",
            f'ZEP准确率({zep_model})': zep_acc,
            'F1分数': f"{stats.get('avg_f1', 0):.4f}",
            '平均检索时间(秒)': f"{stats.get('avg_retrieval_time', 0):.2f}"
        }
        data.append(row)
    
    df = pd.DataFrame(data)
    df.to_excel(writer, sheet_name='问题类型对比-转置', index=False)


def create_individual_question_type_sheets(results: Dict, writer):
    """Create individual sheets for each question type."""
    import pandas as pd
    
    detailed_results = results.get('detailed_results', [])
    
    by_type = {}
    for result in detailed_results:
        qtype = get_question_type(result)
        if qtype not in by_type:
            by_type[qtype] = []
        by_type[qtype].append(result)
    
    sheet_name_map = {
        'single-session-user': '单会话-用户',
        'single-session-assistant': '单会话-助手',
        'single-session-preference': '单会话-偏好',
        'multi-session': '多会话',
        'temporal-reasoning': '时序推理',
        'knowledge-update': '知识更新',
        'unknown': '未知类'
    }
    
    for qtype in sorted(by_type.keys()):
        type_results = by_type[qtype]
        
        data = []
        for result in type_results:
            eval_scores = result.get('evaluation', {}).get('scores', {})
            token_stats = result.get('token_stats', {})
            retrieval = result.get('retrieval_details', {})
            
            data.append({
                'QA索引': result.get('qa_index', 'N/A'),
                '问题': safe_str_truncate(result.get('question', 'N/A'), 60),
                '标准答案': safe_str_truncate(result.get('gold_answer', 'N/A'), 40),
                '生成答案': safe_str_truncate(result.get('generated_answer', 'N/A'), 40),
                'LLM准确': f"{eval_scores.get('llm_accuracy', 0):.2%}",
                'F1': f"{eval_scores.get('token_f1', 0):.4f}",
                'ROUGE-L': f"{eval_scores.get('rougeL_f', 0):.4f}",
                '语义相似度': f"{eval_scores.get('semantic_similarity', 0):.4f}",
                'Token数': token_stats.get('total_tokens', 0),
                '检索时间(s)': f"{retrieval.get('retrieval_time', 0):.2f}",
                '测试时间(s)': f"{result.get('total_time', 0):.2f}"
            })
        
        df = pd.DataFrame(data)
        sheet_name = sheet_name_map.get(qtype, qtype[:15])
        df.to_excel(writer, sheet_name=sheet_name, index=False)


def create_token_analysis_sheet(results: Dict, writer):
    """Create token usage analysis sheet (horizontal layout)."""
    import pandas as pd
    
    detailed_results = results.get('detailed_results', [])
    
    # Overall token statistics
    context_tokens = []
    prompt_tokens = []
    total_tokens = []
    
    by_type = {}
    
    for result in detailed_results:
        token_stats = result.get('token_stats', {})
        qtype = get_question_type(result)
        
        if token_stats:
            ctx = token_stats.get('context_tokens', 0)
            pmt = token_stats.get('prompt_tokens', 0)
            tot = token_stats.get('total_tokens', 0)
            
            context_tokens.append(ctx)
            prompt_tokens.append(pmt)
            total_tokens.append(tot)
            
            if qtype not in by_type:
                by_type[qtype] = {
                    'context': [],
                    'prompt': [],
                    'total': []
                }
            by_type[qtype]['context'].append(ctx)
            by_type[qtype]['prompt'].append(pmt)
            by_type[qtype]['total'].append(tot)
    
    # Create horizontal summary data
    categories = ['全部'] + sorted(by_type.keys())
    
    data = {
        '问题类型': categories,
        '样本数': [],
        '平均上下文Token': [],
        '平均提示Token': [],
        '平均总Token': [],
        '中位数Token': [],
        '最小Token': [],
        '最大Token': [],
        '总Token数': []
    }
    
    # Overall stats
    data['样本数'].append(len(total_tokens))
    data['平均上下文Token'].append(f"{statistics.mean(context_tokens):.0f}" if context_tokens else 0)
    data['平均提示Token'].append(f"{statistics.mean(prompt_tokens):.0f}" if prompt_tokens else 0)
    data['平均总Token'].append(f"{statistics.mean(total_tokens):.0f}" if total_tokens else 0)
    data['中位数Token'].append(f"{statistics.median(total_tokens):.0f}" if total_tokens else 0)
    data['最小Token'].append(min(total_tokens) if total_tokens else 0)
    data['最大Token'].append(max(total_tokens) if total_tokens else 0)
    data['总Token数'].append(sum(total_tokens))
    
    # By type stats
    for qtype in sorted(by_type.keys()):
        tokens = by_type[qtype]
        data['样本数'].append(len(tokens['total']))
        data['平均上下文Token'].append(f"{statistics.mean(tokens['context']):.0f}")
        data['平均提示Token'].append(f"{statistics.mean(tokens['prompt']):.0f}")
        data['平均总Token'].append(f"{statistics.mean(tokens['total']):.0f}")
        data['中位数Token'].append(f"{statistics.median(tokens['total']):.0f}")
        data['最小Token'].append(min(tokens['total']))
        data['最大Token'].append(max(tokens['total']))
        data['总Token数'].append(sum(tokens['total']))
    
    df = pd.DataFrame(data)
    df.to_excel(writer, sheet_name='Token分析', index=False)


def create_cost_estimation_sheet_standalone(results: Dict, writer):
    """Create cost estimation sheet as standalone."""
    import pandas as pd
    
    detailed_results = results.get('detailed_results', [])
    
    # Calculate total tokens
    total_tokens = sum(
        r.get('token_stats', {}).get('total_tokens', 0) 
        for r in detailed_results
    )
    
    pricing = {
        'gpt-4o-mini': 0.15,
        'gpt-4o': 2.50,
        'gpt-4-turbo': 10.00,
        'gpt-3.5-turbo': 0.50,
        'deepseek-chat': 0.14,
        'claude-3-haiku': 0.25,
        'claude-3-sonnet': 3.00
    }
    
    data = []
    for model, price_per_m in pricing.items():
        cost = (total_tokens / 1_000_000) * price_per_m
        data.append({
            '模型': model,
            '价格($/M tokens)': f"{price_per_m:.2f}",
            '总Token数': f"{total_tokens:,}",
            '估算成本($)': f"${cost:.4f}",
            '估算成本(¥)': f"¥{cost * 7.2:.2f}"
        })
    
    df = pd.DataFrame(data)
    df.to_excel(writer, sheet_name='成本估算', index=False)


def create_retrieval_sheet(results: Dict, writer):
    """Create retrieval performance analysis sheet (horizontal layout)."""
    import pandas as pd
    
    detailed_results = results.get('detailed_results', [])
    
    retrieval_times = []
    retrieved_counts = []
    by_type = {}
    
    for result in detailed_results:
        retrieval = result.get('retrieval_details', {})
        qtype = get_question_type(result)
        
        ret_time = retrieval.get('retrieval_time', 0)
        # Entity-relation uses total_entities_retrieved
        ret_count = retrieval.get('total_entities_retrieved', 0)
        
        retrieval_times.append(ret_time)
        retrieved_counts.append(ret_count)
        
        if qtype not in by_type:
            by_type[qtype] = {
                'times': [],
                'counts': []
            }
        by_type[qtype]['times'].append(ret_time)
        by_type[qtype]['counts'].append(ret_count)
    
    categories = ['全部'] + sorted(by_type.keys())
    
    data = {
        '问题类型': categories,
        '样本数': [],
        '平均检索时间(秒)': [],
        '中位数检索时间(秒)': [],
        '最小检索时间(秒)': [],
        '最大检索时间(秒)': [],
        '平均检索数量': []
    }
    
    # Overall
    data['样本数'].append(len(retrieval_times))
    data['平均检索时间(秒)'].append(f"{statistics.mean(retrieval_times):.2f}" if retrieval_times else 0)
    data['中位数检索时间(秒)'].append(f"{statistics.median(retrieval_times):.2f}" if retrieval_times else 0)
    data['最小检索时间(秒)'].append(f"{min(retrieval_times):.2f}" if retrieval_times else 0)
    data['最大检索时间(秒)'].append(f"{max(retrieval_times):.2f}" if retrieval_times else 0)
    data['平均检索数量'].append(f"{statistics.mean(retrieved_counts):.1f}" if retrieved_counts else 0)
    
    # By type
    for qtype in sorted(by_type.keys()):
        metrics = by_type[qtype]
        data['样本数'].append(len(metrics['times']))
        data['平均检索时间(秒)'].append(f"{statistics.mean(metrics['times']):.2f}")
        data['中位数检索时间(秒)'].append(f"{statistics.median(metrics['times']):.2f}")
        data['最小检索时间(秒)'].append(f"{min(metrics['times']):.2f}")
        data['最大检索时间(秒)'].append(f"{max(metrics['times']):.2f}")
        data['平均检索数量'].append(f"{statistics.mean(metrics['counts']):.1f}")
    
    df = pd.DataFrame(data)
    df.to_excel(writer, sheet_name='检索性能', index=False)


def create_error_analysis_sheet(results: Dict, writer):
    """Create error analysis sheet."""
    import pandas as pd
    
    detailed_results = results.get('detailed_results', [])
    
    # Find low accuracy cases
    errors = []
    for result in detailed_results:
        eval_scores = result.get('evaluation', {}).get('scores', {})
        llm_acc = eval_scores.get('llm_accuracy', 0)
        
        if llm_acc < 0.5:  # Consider as error if accuracy < 0.5
            errors.append({
                'QA索引': result.get('qa_index', 'N/A'),
                '问题类型': get_question_type(result),
                '问题': safe_str_truncate(result.get('question', 'N/A'), 80),
                '标准答案': safe_str_truncate(result.get('gold_answer', 'N/A'), 60),
                '生成答案': safe_str_truncate(result.get('generated_answer', 'N/A'), 60),
                'LLM准确率': f"{llm_acc:.2%}",
                'F1分数': f"{eval_scores.get('token_f1', 0):.4f}",
                '检索数量': result.get('retrieval_details', {}).get('total_entities_retrieved', 0),
                'Token数': result.get('token_stats', {}).get('total_tokens', 0)
            })
    
    if errors:
        df = pd.DataFrame(errors)
        df = df.sort_values('LLM准确率')
        df.to_excel(writer, sheet_name='错误分析', index=False)
    else:
        # Create empty sheet with message
        df = pd.DataFrame({'信息': ['没有发现准确率低于50%的案例']})
        df.to_excel(writer, sheet_name='错误分析', index=False)


def create_entity_statistics_sheet(results: Dict, writer):
    """Create entity statistics sheet (specific to entity-relation benchmark)."""
    import pandas as pd
    
    detailed_results = results.get('detailed_results', [])
    
    # Aggregate entity type statistics
    entity_type_counts = defaultdict(int)
    
    for result in detailed_results:
        ret_details = result.get('retrieval_details', {})
        entity_types = ret_details.get('entity_types_found', {})
        
        for etype, count in entity_types.items():
            entity_type_counts[etype] += count
    
    # Create entity type distribution sheet
    entity_dist_data = [
        {'实体类型': etype, '总数': count, '平均每QA': f"{count / len(detailed_results):.2f}"}
        for etype, count in sorted(entity_type_counts.items(), key=lambda x: x[1], reverse=True)
    ]
    
    if entity_dist_data:
        df = pd.DataFrame(entity_dist_data)
    else:
        df = pd.DataFrame({'信息': ['没有实体类型统计数据']})
    
    df.to_excel(writer, sheet_name='实体类型统计', index=False)


def apply_excel_formatting(file_path: str):
    """Apply formatting to Excel file."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = load_workbook(file_path)
        
        # Format each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Header formatting
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    if cell.row > 1:  # Data rows
                        cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                adjusted_width = min(max_length + 3, 60)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze first row
            ws.freeze_panes = 'A2'
            
            # Set row height for header
            ws.row_dimensions[1].height = 25
        
        wb.save(file_path)
        
    except Exception as e:
        print(f"警告: 格式化Excel文件时出错: {e}")


def find_individual_reports_dir(start_path: str = '.') -> Optional[str]:
    """
    Find the most recent individual_reports directory for entity_relation.
    """
    start = Path(start_path)
    
    # Search patterns
    search_patterns = [
        start / 'results' / 'entity_relation_no_expand',
        start / 'benchmark_longmemeval' / 'task_eval' / 'results' / 'entity_relation_no_expand',
        Path('.') / 'benchmark_longmemeval' / 'task_eval' / 'results' / 'entity_relation_no_expand',
    ]
    
    for results_base in search_patterns:
        if results_base.exists():
            # Find all run_* directories
            run_dirs = sorted(results_base.glob('run_*'), reverse=True)
            
            for run_dir in run_dirs:
                reports_dir = run_dir / 'individual_reports'
                if reports_dir.exists() and list(reports_dir.glob('qa_*.json')):
                    return str(reports_dir)
    
    return None


def main():
    """Main function."""
    parser = argparse.ArgumentParser(
        description='生成Entity-Relation基准测试的Excel分析报告（横版布局，ZEP风格）'
    )
    parser.add_argument(
        '--input_dir',
        type=str,
        nargs='?',
        help='individual_reports目录路径(可选，默认自动查找)'
    )
    parser.add_argument(
        '-o', '--output',
        type=str,
        default=None,
        help='输出Excel文件路径(可选，默认与输入同目录)'
    )
    parser.add_argument(
        '--zep-model',
        type=str,
        default='gpt-4o-mini',
        choices=['gpt-4o-mini', 'gpt-4o'],
        help='ZEP对比使用的模型 (默认: gpt-4o-mini)'
    )
    
    args = parser.parse_args()
    
    # Find input directory
    if args.input_dir:
        input_dir = args.input_dir
    else:
        print(" 自动查找最新的entity_relation individual_reports目录...")
        input_dir = find_individual_reports_dir()
        if not input_dir:
            print("错误: 未找到individual_reports目录，请使用 --input_dir 指定")
            return
        print(f" 找到目录: {input_dir}")
    
    if not os.path.exists(input_dir):
        print(f"错误: 目录不存在: {input_dir}")
        return
    
    # Determine output path
    if args.output:
        output_path = args.output
    else:
        input_path = Path(input_dir)
        parent_dir = input_path.parent
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = parent_dir / f"entity_relation_analysis_{timestamp}.xlsx"
    
    print(f"正在加载benchmark结果: {input_dir}")
    
    try:
        detailed_results, summary = load_individual_results(input_dir)
    except Exception as e:
        print(f"错误: {e}")
        return
    
    print(f"\n统计摘要:")
    print(f"  总测试数: {summary.get('total_tests', 0)}")
    print(f"  成功数: {summary.get('successful_tests', 0)}")
    print(f"  LLM准确率: {summary.get('overall_llm_accuracy', 0):.2%}")
    print(f"  平均F1: {summary.get('overall_f1', 0):.4f}")
    
    print(f"\n正在生成Excel报告: {output_path}")
    print(f"ZEP对比模型: {args.zep_model}")
    success = create_excel_report(detailed_results, summary, str(output_path), zep_model=args.zep_model)
    
    if success:
        print(f"\n Excel分析报告已生成: {output_path}")
        print(f"   ZEP基准: {args.zep_model}")
        print("\n报告包含以下工作表:")
        print("  1. 总体概览 - 基准测试总体统计（含ZEP对比）")
        print("  2. 问题类型对比 - ZEP风格横向对比表（含ZEP准确率）")
        print("  3. 问题类型对比-转置 - 转置版本（便于截图到Word）")
        print("  4+. 各问题类型独立详细数据表")
        print("  Token分析 - Token使用统计（横版）")
        print("  成本估算 - 不同模型成本估算")
        print("  检索性能 - 检索时间分析（横版）")
        print("  错误分析 - 低准确率案例")
        print("  实体类型统计 - Entity-Relation特有")
    else:
        print("\n 生成报告失败")


if __name__ == '__main__':
    main()
