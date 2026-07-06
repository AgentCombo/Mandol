#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate Excel Analysis Report for Triple Retrieval Fusion Benchmark Results

This script reads all qa_*_report.json files from the individual_reports directory
of triple fusion benchmark, aggregates the results, and creates a comprehensive
Excel report with multi-level retrieval analysis.

Features:
- Triple retrieval fusion analysis (Sentence + Episodic + Entity)
- Question type breakdown
- Token usage and cost estimation
- Retrieval performance metrics
- EverMemOS baseline comparison
"""

import json
import os
import re
import statistics
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from datetime import datetime
import argparse

EXCEL_SAFE_PATH_LIMIT = 240


def estimate_windows_path_length(path: Path) -> int:
    """Estimate the path length Excel may see on Windows/WSL."""
    resolved = path.resolve()
    linux_path = str(resolved)
    wsl_distro = os.environ.get("WSL_DISTRO_NAME", "Ubuntu")
    wsl_unc = f"\\\\wsl.localhost\\{wsl_distro}" + linux_path.replace("/", "\\")
    return max(len(linux_path), len(wsl_unc))


def build_default_output_path(reports_path: Path, timestamp: str) -> Path:
    """Build a default Excel output path that avoids Windows Excel path limits."""
    filename = f"triple_fusion_report_{timestamp}.xlsx"
    legacy_path = reports_path.parent / f"triple_fusion_analysis_report_{timestamp}.xlsx"

    if estimate_windows_path_length(legacy_path) <= EXCEL_SAFE_PATH_LIMIT:
        return legacy_path

    short_dir = Path.cwd() / "excel_reports"
    return short_dir / filename


def warn_if_excel_path_long(output_path: Path, is_explicit: bool) -> None:
    """Warn when Excel may refuse to open the generated file due to path length."""
    estimated_len = estimate_windows_path_length(output_path)
    if estimated_len <= EXCEL_SAFE_PATH_LIMIT:
        return

    source = "--output 指定的路径" if is_explicit else "默认输出路径"
    print(
        f"  警告: {source}较长，Windows Excel 可能无法打开 "
        f"(估算长度 {estimated_len}, 建议 <= {EXCEL_SAFE_PATH_LIMIT})"
    )
    print("   建议使用 --output excel_reports/report.xlsx 或移动到更短目录。")



# Data Loading & Processing


def load_individual_results(individual_reports_dir: str) -> Tuple[List[Dict], Dict]:
    """
    Load all qa_*_report.json files from the individual_reports directory.
    
    Returns:
        Tuple of (detailed_results list, computed summary dict)
    """
    reports_path = Path(individual_reports_dir)
    
    if not reports_path.exists():
        raise FileNotFoundError(f"目录不存在: {individual_reports_dir}")
    
    # Find all qa_*_report.json files
    json_files = list(reports_path.glob('qa_*_report.json'))
    
    if not json_files:
        raise FileNotFoundError(f"在 {individual_reports_dir} 中未找到 qa_*_report.json 文件")
    
    # Sort by qa number
    def extract_qa_number(filepath):
        match = re.search(r'qa_(\d+)_report\.json', filepath.name)
        return int(match.group(1)) if match else 0
    
    json_files = sorted(json_files, key=extract_qa_number)
    
    detailed_results = []
    failed_count = 0
    
    print(f"正在加载 {len(json_files)} 个QA结果文件...")
    
    for json_file in json_files:
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                result = json.load(f)
                detailed_results.append(result)
        except Exception as e:
            print(f"警告: 读取 {json_file} 失败: {e}")
            failed_count += 1
    
    print(f"成功加载: {len(detailed_results)} 个, 失败: {failed_count} 个")
    
    # Compute summary from detailed results
    summary = compute_summary(detailed_results)
    
    return detailed_results, summary


def get_question_type(result: Dict) -> str:
    """Helper to extract question type from the JSON structure."""
    qtype = result.get('scores', {}).get('input_info', {}).get('question_type')
    if not qtype:
        
        qtype = result.get('question_type', 'unknown')
    return qtype or 'unknown'


def compute_summary(detailed_results: List[Dict]) -> Dict:
    """
    Compute summary statistics from detailed results.
    """
    if not detailed_results:
        return {}
    
    total_tests = len(detailed_results)
    successful_tests = sum(1 for r in detailed_results if r.get('success', False))
    failed_tests = total_tests - successful_tests
    
    # Collect metrics
    llm_accuracies = []
    f1_scores = []
    retrieval_times = []
    
    # Triple retrieval specific metrics
    sentence_retrieval_counts = []
    episodic_retrieval_counts = []
    entity_retrieval_counts = []
    
    # By question type
    by_question_type = defaultdict(lambda: {
        'count': 0,
        'llm_accuracies': [],
        'f1_scores': [],
        'retrieval_times': [],
        'sentence_counts': [],
        'episodic_counts': [],
        'entity_counts': []
    })
    
    # Get retrieval config from first result
    retrieval_config = {}
    if detailed_results:
        first_result = detailed_results[0]
        retrieval_details = first_result.get('retrieval_details', {})
        retrieval_config = {
            'sentence_enabled': retrieval_details.get('sentence_enabled', True),
            'episodic_enabled': retrieval_details.get('episodic_enabled', True),
            'entity_enabled': retrieval_details.get('entity_enabled', True),
            'rerank_method': retrieval_details.get('rerank_method', 'baai'),
            'fusion_method': retrieval_details.get('fusion_method', 'concatenation')
        }
    
    for result in detailed_results:
        # Skip failed results
        if not result.get('success', False):
            continue
        
        # Get scores
        scores = result.get('scores', {}).get('scores', {})
        llm_acc = scores.get('llm_accuracy', 0.0)
        f1 = scores.get('token_f1', 0.0)
        
        # Get retrieval details
        retrieval_details = result.get('retrieval_details', {})
        ret_time = retrieval_details.get('total_retrieval_time', 0.0)
        
        sentence_count = retrieval_details.get('sentence_retrieved_count', 0)
        episodic_count = retrieval_details.get('episodic_retrieved_count', 0)
        entity_count = retrieval_details.get('entity_retrieved_count', 0)
        
        llm_accuracies.append(llm_acc)
        f1_scores.append(f1)
        retrieval_times.append(ret_time)
        
        sentence_retrieval_counts.append(sentence_count)
        episodic_retrieval_counts.append(episodic_count)
        entity_retrieval_counts.append(entity_count)
        
        # By question type
        qtype = get_question_type(result)
        by_question_type[qtype]['count'] += 1
        by_question_type[qtype]['llm_accuracies'].append(llm_acc)
        by_question_type[qtype]['f1_scores'].append(f1)
        by_question_type[qtype]['retrieval_times'].append(ret_time)
        by_question_type[qtype]['sentence_counts'].append(sentence_count)
        by_question_type[qtype]['episodic_counts'].append(episodic_count)
        by_question_type[qtype]['entity_counts'].append(entity_count)
    
    # Compute averages for each question type
    by_type_summary = {}
    for qtype, data in by_question_type.items():
        by_type_summary[qtype] = {
            'count': data['count'],
            'avg_llm_accuracy': statistics.mean(data['llm_accuracies']) if data['llm_accuracies'] else 0,
            'avg_f1': statistics.mean(data['f1_scores']) if data['f1_scores'] else 0,
            'avg_retrieval_time': statistics.mean(data['retrieval_times']) if data['retrieval_times'] else 0,
            'avg_sentence_count': statistics.mean(data['sentence_counts']) if data['sentence_counts'] else 0,
            'avg_episodic_count': statistics.mean(data['episodic_counts']) if data['episodic_counts'] else 0,
            'avg_entity_count': statistics.mean(data['entity_counts']) if data['entity_counts'] else 0
        }
    
    # Build summary
    summary = {
        'baseline_version': 'triple_fusion',
        'dataset_size': f'{total_tests} QAs',
        'total_tests': total_tests,
        'successful_tests': successful_tests,
        'failed_tests': failed_tests,
        'overall_llm_accuracy': statistics.mean(llm_accuracies) if llm_accuracies else 0,
        'overall_f1': statistics.mean(f1_scores) if f1_scores else 0,
        'avg_retrieval_time': statistics.mean(retrieval_times) if retrieval_times else 0,
        'avg_sentence_count': statistics.mean(sentence_retrieval_counts) if sentence_retrieval_counts else 0,
        'avg_episodic_count': statistics.mean(episodic_retrieval_counts) if episodic_retrieval_counts else 0,
        'avg_entity_count': statistics.mean(entity_retrieval_counts) if entity_retrieval_counts else 0,
        'retrieval_config': retrieval_config,
        'by_question_type': by_type_summary
    }
    
    return summary


def safe_str_truncate(value, max_len: int = 50) -> str:
    """Safely convert value to string and truncate if needed."""
    if value is None:
        return ""
    str_value = str(value)
    if len(str_value) > max_len:
        return str_value[:max_len] + "..."
    return str_value


def get_evermemos_accuracy(question_type: str) -> str:
    """Get EverMemOS accuracy for specific question type (LongMemEval dataset)."""
    evermemos_data = {
        'overall': 0.8300,
        'single-session-user': 0.9714,
        'single-session-assistant': 0.8571,
        'single-session-preference': 0.9333,
        'multi-session': 0.7368,
        'knowledge-update': 0.8974,
        'temporal-reasoning': 0.7744
    }
    
    accuracy = evermemos_data.get(question_type, None)
    if accuracy is not None:
        return f"{accuracy:.2%}"
    return 'N/A'



# Excel Report Generation


def create_excel_report(detailed_results: List[Dict], summary: Dict, output_path: str):
    """Create Excel report with multiple analysis sheets."""
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("错误: 需要安装 pandas 和 openpyxl")
        print("请运行: pip install pandas openpyxl")
        return False
    
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    # Create Excel writer
    writer = pd.ExcelWriter(output_path, engine='openpyxl')
    
    # Prepare results dict for compatibility
    results = {
        'summary': summary,
        'detailed_results': detailed_results
    }
    
    # Sheet 1: Overall Summary
    create_summary_sheet_horizontal(results, writer)
    
    # Sheet 2: Question Type Comparison
    create_question_type_comparison_sheet(results, writer)
    
    # Sheet 3: Question Type Comparison Transposed
    create_question_type_comparison_transposed(results, writer)
    
    # Sheets 4+: Individual Question Type Details
    create_individual_question_type_sheets(results, writer)
    
    # Triple Retrieval Analysis
    create_triple_retrieval_analysis_sheet(results, writer)
    
    # Token Usage Analysis
    create_token_analysis_sheet(results, writer)
    
    # Cost Estimation
    create_cost_estimation_sheet_standalone(results, writer)
    
    # Retrieval Performance
    create_retrieval_sheet(results, writer)
    
    # Error Analysis
    create_error_analysis_sheet(results, writer)
    
    # All Results Detail
    create_all_results_detail_sheet(results, writer)
    
    # Save and close
    writer.close()
    
    # Apply formatting
    apply_excel_formatting(output_path)
    
    print(f"\n Excel报告已生成: {output_path}")
    return True


def create_summary_sheet_horizontal(results: Dict, writer):
    """Create overall summary sheet in horizontal layout with EverMemOS comparison."""
    import pandas as pd
    
    summary = results.get('summary', {})
    evermemos_overall = get_evermemos_accuracy('overall')
    
    data = {
        '基线版本': [summary.get('baseline_version', 'N/A')],
        '数据集': [summary.get('dataset_size', 'N/A')],
        '总测试数': [summary.get('total_tests', 0)],
        '成功数': [summary.get('successful_tests', 0)],
        '失败数': [summary.get('failed_tests', 0)],
        'LLM准确率': [f"{summary.get('overall_llm_accuracy', 0):.2%}"],
        'EverMemOS准确率': [evermemos_overall],
        '平均F1': [f"{summary.get('overall_f1', 0):.4f}"],
        '平均检索时间(秒)': [f"{summary.get('avg_retrieval_time', 0):.2f}"],
        '平均Sentence检索数': [f"{summary.get('avg_sentence_count', 0):.1f}"],
        '平均Episodic检索数': [f"{summary.get('avg_episodic_count', 0):.1f}"],
        '平均Entity检索数': [f"{summary.get('avg_entity_count', 0):.1f}"],
        '融合方法': [summary.get('retrieval_config', {}).get('fusion_method', 'N/A')],
        '重排方法': [summary.get('retrieval_config', {}).get('rerank_method', 'N/A')]
    }
    
    df = pd.DataFrame(data)
    df.to_excel(writer, sheet_name='总体概览', index=False)


def create_question_type_comparison_sheet(results: Dict, writer):
    """Create question type comparison sheet with EverMemOS accuracy."""
    import pandas as pd
    
    by_type = results.get('summary', {}).get('by_question_type', {})
    question_types = sorted(by_type.keys())
    
    data = {
        '指标': ['样本数', 'LLM准确率', 'EverMemOS准确率', 'F1分数', '平均检索时间(秒)', 
                'Sentence检索数', 'Episodic检索数', 'Entity检索数']
    }
    
    for qtype in question_types:
        type_data = by_type[qtype]
        evermemos_acc = get_evermemos_accuracy(qtype)
        
        data[qtype] = [
            type_data['count'],
            f"{type_data['avg_llm_accuracy']:.2%}",
            evermemos_acc,
            f"{type_data['avg_f1']:.4f}",
            f"{type_data['avg_retrieval_time']:.2f}",
            f"{type_data.get('avg_sentence_count', 0):.1f}",
            f"{type_data.get('avg_episodic_count', 0):.1f}",
            f"{type_data.get('avg_entity_count', 0):.1f}"
        ]
    
    df = pd.DataFrame(data)
    df.to_excel(writer, sheet_name='问题类型对比', index=False)


def create_question_type_comparison_transposed(results: Dict, writer):
    """Create transposed question type comparison sheet."""
    import pandas as pd
    
    by_type = results.get('summary', {}).get('by_question_type', {})
    question_types = sorted(by_type.keys())
    
    data = []
    
    for qtype in question_types:
        type_data = by_type[qtype]
        evermemos_acc = get_evermemos_accuracy(qtype)
        
        row = {
            '问题类型': qtype,
            '样本数': type_data['count'],
            'LLM准确率': f"{type_data['avg_llm_accuracy']:.2%}",
            'EverMemOS准确率': evermemos_acc,
            'F1分数': f"{type_data['avg_f1']:.4f}",
            '平均检索时间(秒)': f"{type_data['avg_retrieval_time']:.2f}",
            'Sentence检索数': f"{type_data.get('avg_sentence_count', 0):.1f}",
            'Episodic检索数': f"{type_data.get('avg_episodic_count', 0):.1f}",
            'Entity检索数': f"{type_data.get('avg_entity_count', 0):.1f}"
        }
        data.append(row)
    
    df = pd.DataFrame(data)
    df.to_excel(writer, sheet_name='问题类型对比-转置', index=False)


def create_individual_question_type_sheets(results: Dict, writer):
    """Create individual sheets for each question type."""
    import pandas as pd
    
    detailed_results = results.get('detailed_results', [])
    
    # Group by question type
    by_type = {}
    for result in detailed_results:
        qtype = get_question_type(result)
        if qtype not in by_type:
            by_type[qtype] = []
        by_type[qtype].append(result)
    
    # Create sheet for each type
    for qtype in sorted(by_type.keys()):
        results_for_type = by_type[qtype]
        
        data = []
        for result in results_for_type:
            scores = result.get('scores', {}).get('scores', {})
            retrieval_details = result.get('retrieval_details', {})
            token_stats = result.get('token_stats', {})
            
            row = {
                'QA索引': result.get('qa_index', 'N/A'),
                '问题': safe_str_truncate(result.get('question', ''), 60),
                '标准答案': safe_str_truncate(result.get('ground_truth', ''), 40),
                '生成答案': safe_str_truncate(result.get('generated_answer', ''), 40),
                'LLM准确率': scores.get('llm_accuracy', 0.0),
                'F1分数': scores.get('token_f1', 0.0),
                'Sentence检索数': retrieval_details.get('sentence_retrieved_count', 0),
                'Episodic检索数': retrieval_details.get('episodic_retrieved_count', 0),
                'Entity检索数': retrieval_details.get('entity_retrieved_count', 0),
                '检索时间(秒)': retrieval_details.get('total_retrieval_time', 0.0),
                'Prompt Tokens': token_stats.get('prompt_tokens', 0),
                '成功': result.get('success', False)
            }
            data.append(row)
        
        df = pd.DataFrame(data)
        # Truncate sheet name to 31 chars (Excel limit)
        sheet_name = f"类型_{qtype}"[:31]
        df.to_excel(writer, sheet_name=sheet_name, index=False)


def create_triple_retrieval_analysis_sheet(results: Dict, writer):
    """Create triple retrieval fusion analysis sheet."""
    import pandas as pd
    
    detailed_results = results.get('detailed_results', [])
    
    data = []
    for result in detailed_results:
        if not result.get('success', False):
            continue
        
        retrieval_details = result.get('retrieval_details', {})
        scores = result.get('scores', {}).get('scores', {})
        
        row = {
            'QA索引': result.get('qa_index', 'N/A'),
            '问题类型': get_question_type(result),
            'Sentence启用': '' if retrieval_details.get('sentence_enabled', False) else '',
            'Sentence检索数': retrieval_details.get('sentence_retrieved_count', 0),
            'Sentence时间(秒)': retrieval_details.get('sentence_retrieval_time', 0.0),
            'Episodic启用': '' if retrieval_details.get('episodic_enabled', False) else '',
            'Episodic检索数': retrieval_details.get('episodic_retrieved_count', 0),
            'Episodic时间(秒)': retrieval_details.get('episodic_retrieval_time', 0.0),
            'Entity启用': '' if retrieval_details.get('entity_enabled', False) else '',
            'Entity检索数': retrieval_details.get('entity_retrieved_count', 0),
            'Entity时间(秒)': retrieval_details.get('entity_retrieval_time', 0.0),
            '总检索时间(秒)': retrieval_details.get('total_retrieval_time', 0.0),
            'LLM准确率': scores.get('llm_accuracy', 0.0),
            'F1分数': scores.get('token_f1', 0.0)
        }
        data.append(row)
    
    df = pd.DataFrame(data)
    df.to_excel(writer, sheet_name='三重检索分析', index=False)


def create_token_analysis_sheet(results: Dict, writer):
    """Create token usage analysis sheet."""
    import pandas as pd
    
    detailed_results = results.get('detailed_results', [])
    
    # Overall token statistics
    prompt_tokens = []
    completion_tokens = []
    total_tokens = []
    sentence_tokens = []
    episodic_tokens = []
    entity_tokens = []
    
    by_type = {}
    
    for result in detailed_results:
        if not result.get('success', False):
            continue
        
        token_stats = result.get('token_stats', {})
        prompt_tokens.append(token_stats.get('prompt_tokens', 0))
        completion_tokens.append(token_stats.get('completion_tokens', 0))
        total_tokens.append(token_stats.get('total_tokens', 0))
        sentence_tokens.append(token_stats.get('sentence_context_tokens', 0))
        episodic_tokens.append(token_stats.get('episodic_context_tokens', 0))
        entity_tokens.append(token_stats.get('entity_context_tokens', 0))
        
        qtype = get_question_type(result)
        if qtype not in by_type:
            by_type[qtype] = {
                'prompt': [], 'completion': [], 'total': [],
                'sentence': [], 'episodic': [], 'entity': []
            }
        by_type[qtype]['prompt'].append(token_stats.get('prompt_tokens', 0))
        by_type[qtype]['completion'].append(token_stats.get('completion_tokens', 0))
        by_type[qtype]['total'].append(token_stats.get('total_tokens', 0))
        by_type[qtype]['sentence'].append(token_stats.get('sentence_context_tokens', 0))
        by_type[qtype]['episodic'].append(token_stats.get('episodic_context_tokens', 0))
        by_type[qtype]['entity'].append(token_stats.get('entity_context_tokens', 0))
    
    # Create summary data
    categories = ['全部'] + sorted(by_type.keys())
    
    data = {
        '问题类型': categories,
        '样本数': [],
        '平均Prompt': [],
        '平均Completion': [],
        '平均Total': [],
        '平均Sentence上下文': [],
        '平均Episodic上下文': [],
        '平均Entity上下文': [],
        '总Token数': []
    }
    
    # Overall stats
    data['样本数'].append(len(prompt_tokens))
    data['平均Prompt'].append(f"{statistics.mean(prompt_tokens):.0f}" if prompt_tokens else 0)
    data['平均Completion'].append(f"{statistics.mean(completion_tokens):.0f}" if completion_tokens else 0)
    data['平均Total'].append(f"{statistics.mean(total_tokens):.0f}" if total_tokens else 0)
    data['平均Sentence上下文'].append(f"{statistics.mean(sentence_tokens):.0f}" if sentence_tokens else 0)
    data['平均Episodic上下文'].append(f"{statistics.mean(episodic_tokens):.0f}" if episodic_tokens else 0)
    data['平均Entity上下文'].append(f"{statistics.mean(entity_tokens):.0f}" if entity_tokens else 0)
    data['总Token数'].append(sum(total_tokens))
    
    # By type stats
    for qtype in sorted(by_type.keys()):
        type_data = by_type[qtype]
        data['样本数'].append(len(type_data['total']))
        data['平均Prompt'].append(f"{statistics.mean(type_data['prompt']):.0f}" if type_data['prompt'] else 0)
        data['平均Completion'].append(f"{statistics.mean(type_data['completion']):.0f}" if type_data['completion'] else 0)
        data['平均Total'].append(f"{statistics.mean(type_data['total']):.0f}" if type_data['total'] else 0)
        data['平均Sentence上下文'].append(f"{statistics.mean(type_data['sentence']):.0f}" if type_data['sentence'] else 0)
        data['平均Episodic上下文'].append(f"{statistics.mean(type_data['episodic']):.0f}" if type_data['episodic'] else 0)
        data['平均Entity上下文'].append(f"{statistics.mean(type_data['entity']):.0f}" if type_data['entity'] else 0)
        data['总Token数'].append(sum(type_data['total']))
    
    df = pd.DataFrame(data)
    df.to_excel(writer, sheet_name='Token分析', index=False)


def create_cost_estimation_sheet_standalone(results: Dict, writer):
    """Create cost estimation sheet."""
    import pandas as pd
    
    detailed_results = results.get('detailed_results', [])
    
    # Calculate total tokens
    total_prompt = sum(r.get('token_stats', {}).get('prompt_tokens', 0) for r in detailed_results)
    total_completion = sum(r.get('token_stats', {}).get('completion_tokens', 0) for r in detailed_results)
    
    # Pricing (per 1M tokens)
    pricing = {
        'gpt-4o-mini': {'input': 0.15, 'output': 0.60},
        'gpt-4o': {'input': 2.50, 'output': 10.00},
        'gpt-4-turbo': {'input': 10.00, 'output': 30.00},
        'deepseek-chat': {'input': 0.14, 'output': 0.28},
        'claude-3-haiku': {'input': 0.25, 'output': 1.25},
        'claude-3-sonnet': {'input': 3.00, 'output': 15.00}
    }
    
    data = []
    for model, prices in pricing.items():
        input_cost = (total_prompt / 1_000_000) * prices['input']
        output_cost = (total_completion / 1_000_000) * prices['output']
        total_cost = input_cost + output_cost
        
        row = {
            '模型': model,
            'Input价格($/1M)': f"${prices['input']:.2f}",
            'Output价格($/1M)': f"${prices['output']:.2f}",
            'Input成本($)': f"${input_cost:.4f}",
            'Output成本($)': f"${output_cost:.4f}",
            '总成本($)': f"${total_cost:.4f}"
        }
        data.append(row)
    
    df = pd.DataFrame(data)
    df.to_excel(writer, sheet_name='成本估算', index=False)


def create_retrieval_sheet(results: Dict, writer):
    """Create retrieval performance analysis sheet."""
    import pandas as pd
    
    detailed_results = results.get('detailed_results', [])
    
    retrieval_times = []
    sentence_counts = []
    episodic_counts = []
    entity_counts = []
    
    by_type = {}
    
    for result in detailed_results:
        if not result.get('success', False):
            continue
        
        retrieval_details = result.get('retrieval_details', {})
        ret_time = retrieval_details.get('total_retrieval_time', 0.0)
        
        retrieval_times.append(ret_time)
        sentence_counts.append(retrieval_details.get('sentence_retrieved_count', 0))
        episodic_counts.append(retrieval_details.get('episodic_retrieved_count', 0))
        entity_counts.append(retrieval_details.get('entity_retrieved_count', 0))
        
        qtype = get_question_type(result)
        if qtype not in by_type:
            by_type[qtype] = {
                'times': [], 'sentence': [], 'episodic': [], 'entity': []
            }
        by_type[qtype]['times'].append(ret_time)
        by_type[qtype]['sentence'].append(retrieval_details.get('sentence_retrieved_count', 0))
        by_type[qtype]['episodic'].append(retrieval_details.get('episodic_retrieved_count', 0))
        by_type[qtype]['entity'].append(retrieval_details.get('entity_retrieved_count', 0))
    
    categories = ['全部'] + sorted(by_type.keys())
    
    data = {
        '问题类型': categories,
        '样本数': [],
        '平均检索时间(秒)': [],
        '中位数时间(秒)': [],
        '平均Sentence数': [],
        '平均Episodic数': [],
        '平均Entity数': []
    }
    
    # Overall
    data['样本数'].append(len(retrieval_times))
    data['平均检索时间(秒)'].append(f"{statistics.mean(retrieval_times):.2f}" if retrieval_times else 0)
    data['中位数时间(秒)'].append(f"{statistics.median(retrieval_times):.2f}" if retrieval_times else 0)
    data['平均Sentence数'].append(f"{statistics.mean(sentence_counts):.1f}" if sentence_counts else 0)
    data['平均Episodic数'].append(f"{statistics.mean(episodic_counts):.1f}" if episodic_counts else 0)
    data['平均Entity数'].append(f"{statistics.mean(entity_counts):.1f}" if entity_counts else 0)
    
    # By type
    for qtype in sorted(by_type.keys()):
        type_data = by_type[qtype]
        data['样本数'].append(len(type_data['times']))
        data['平均检索时间(秒)'].append(f"{statistics.mean(type_data['times']):.2f}" if type_data['times'] else 0)
        data['中位数时间(秒)'].append(f"{statistics.median(type_data['times']):.2f}" if type_data['times'] else 0)
        data['平均Sentence数'].append(f"{statistics.mean(type_data['sentence']):.1f}" if type_data['sentence'] else 0)
        data['平均Episodic数'].append(f"{statistics.mean(type_data['episodic']):.1f}" if type_data['episodic'] else 0)
        data['平均Entity数'].append(f"{statistics.mean(type_data['entity']):.1f}" if type_data['entity'] else 0)
    
    df = pd.DataFrame(data)
    df.to_excel(writer, sheet_name='检索性能', index=False)


def create_error_analysis_sheet(results: Dict, writer):
    """Create error analysis sheet."""
    import pandas as pd
    
    detailed_results = results.get('detailed_results', [])
    
    failed_results = [r for r in detailed_results if not r.get('success', False)]
    
    data = []
    for result in failed_results:
        row = {
            'QA索引': result.get('qa_index', 'N/A'),
            '问题': safe_str_truncate(result.get('question', ''), 80),
            '问题类型': get_question_type(result),
            '错误信息': safe_str_truncate(result.get('error_message', ''), 100)
        }
        data.append(row)
    
    if data:
        df = pd.DataFrame(data)
    else:
        df = pd.DataFrame({'说明': [' 没有失败的测试用例']})
    
    df.to_excel(writer, sheet_name='错误分析', index=False)


def create_all_results_detail_sheet(results: Dict, writer):
    """Create all results detail sheet."""
    import pandas as pd
    
    detailed_results = results.get('detailed_results', [])
    
    data = []
    for result in detailed_results:
        scores = result.get('scores', {}).get('scores', {})
        retrieval_details = result.get('retrieval_details', {})
        token_stats = result.get('token_stats', {})
        
        row = {
            'QA索引': result.get('qa_index', 'N/A'),
            '问题类型': get_question_type(result),
            '问题': safe_str_truncate(result.get('question', ''), 60),
            '标准答案': safe_str_truncate(result.get('ground_truth', ''), 40),
            '生成答案': safe_str_truncate(result.get('generated_answer', ''), 40),
            'LLM准确率': scores.get('llm_accuracy', 0.0),
            'F1分数': scores.get('token_f1', 0.0),
            'ROUGE-L': scores.get('rougeL_f', 0.0),
            'Sentence数': retrieval_details.get('sentence_retrieved_count', 0),
            'Episodic数': retrieval_details.get('episodic_retrieved_count', 0),
            'Entity数': retrieval_details.get('entity_retrieved_count', 0),
            '检索时间': retrieval_details.get('total_retrieval_time', 0.0),
            'Prompt Tokens': token_stats.get('prompt_tokens', 0),
            'Total Tokens': token_stats.get('total_tokens', 0),
            '成功': result.get('success', False)
        }
        data.append(row)
    
    df = pd.DataFrame(data)
    df.to_excel(writer, sheet_name='所有结果详情', index=False)


def apply_excel_formatting(file_path: str):
    """Apply formatting to Excel file."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return
    
    wb = load_workbook(file_path)
    
    # Header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Data cell alignment style
    data_alignment = Alignment(horizontal="center", vertical="center")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Format header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Format data cells (row 2 onwards) with center alignment
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.alignment = data_alignment
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    wb.save(file_path)



# Main Execution


def find_individual_reports_dir(start_path: str = '.') -> Optional[str]:
    """
    Find individual_reports directory in the given path.
    
    Search order:
    1. start_path/individual_reports
    2. start_path/results/*/individual_reports
    3. start_path/../individual_reports
    """
    start = Path(start_path).resolve()
    
    # Direct check
    direct_path = start / 'individual_reports'
    if direct_path.exists() and direct_path.is_dir():
        return str(direct_path)
    
    # Check in results subdirectories
    results_dir = start / 'results'
    if results_dir.exists():
        for subdir in results_dir.iterdir():
            if subdir.is_dir():
                reports_dir = subdir / 'individual_reports'
                if reports_dir.exists():
                    return str(reports_dir)
    
    # Check parent directory
    parent_reports = start.parent / 'individual_reports'
    if parent_reports.exists() and parent_reports.is_dir():
        return str(parent_reports)
    
    return None


def main():
    parser = argparse.ArgumentParser(
        description="Generate Excel Analysis Report for Triple Retrieval Fusion Benchmark"
    )
    
    parser.add_argument(
        '--reports-dir',
        type=str,
        default=None,
        help='Path to individual_reports directory containing qa_*_report.json files'
    )
    
    parser.add_argument(
        '--output',
        type=str,
        default=None,
        help='Output Excel file path (default: short excel_reports/ path if reports-dir is too long for Windows Excel)'
    )
    
    args = parser.parse_args()
    
    # Determine reports directory
    if args.reports_dir:
        reports_dir = args.reports_dir
    else:
        print("未指定 --reports-dir，正在自动搜索...")
        reports_dir = find_individual_reports_dir()
        if not reports_dir:
            print(" 错误: 未找到 individual_reports 目录")
            print("请使用 --reports-dir 参数指定路径")
            return 1
        print(f" 找到报告目录: {reports_dir}")
    
    reports_path = Path(reports_dir)
    if not reports_path.exists():
        print(f" 错误: 目录不存在: {reports_dir}")
        return 1
    
    # Determine output path
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    if args.output:
        output_path = Path(args.output)
    else:
        output_path = build_default_output_path(reports_path, timestamp)
        legacy_path = reports_path.parent / f'triple_fusion_analysis_report_{timestamp}.xlsx'
        if output_path != legacy_path:
            print("  默认输出目录路径过长，已自动改写到较短路径，避免 Windows Excel 打不开。")

    warn_if_excel_path_long(output_path, is_explicit=bool(args.output))
    
    print("\n" + "="*80)
    print(" Triple Retrieval Fusion Excel 报告生成器")
    print("="*80)
    print(f" 输入目录: {reports_dir}")
    print(f" 输出文件: {output_path}")
    print("="*80 + "\n")
    
    try:
        # Load data
        detailed_results, summary = load_individual_results(reports_dir)
        
        print(f"\n 数据统计:")
        print(f"   总测试数: {summary.get('total_tests', 0)}")
        print(f"   成功数: {summary.get('successful_tests', 0)}")
        print(f"   失败数: {summary.get('failed_tests', 0)}")
        print(f"   LLM准确率: {summary.get('overall_llm_accuracy', 0):.2%}")
        print(f"   平均F1: {summary.get('overall_f1', 0):.4f}")
        
        # Generate Excel report
        print(f"\n 正在生成Excel报告...")
        success = create_excel_report(detailed_results, summary, str(output_path))
        
        if success:
            print("\n" + "="*80)
            print(" 报告生成成功!")
            print(f" 文件位置: {output_path}")
            print("="*80)
            return 0
        else:
            print("\n 报告生成失败")
            return 1
            
    except Exception as e:
        print(f"\n 错误: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == '__main__':
    exit(main())
